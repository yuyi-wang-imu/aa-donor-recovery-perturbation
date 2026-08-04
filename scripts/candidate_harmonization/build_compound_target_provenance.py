#!/usr/bin/env python3
"""Normalize the archived compound and compound-target provenance tables.

The compound workbook is the archived TCMSP-derived active-compound table.
The target workbook is the archived SwissTargetPrediction output after the
original study's target-name standardization.  This script does not query
either database and does not impute missing compound-target predictions.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


def normalize(value: object) -> str:
    return "" if value is None else str(value).strip()


def read_rows(path: Path) -> list[dict[str, str]]:
    worksheet = load_workbook(path, read_only=True, data_only=True).worksheets[0]
    header = [normalize(cell.value) for cell in worksheet[1]]
    rows: list[dict[str, str]] = []
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        row = {key: normalize(value) for key, value in zip(header, values, strict=False)}
        if any(row.values()):
            rows.append(row)
    return rows


def write_csv(path: Path, rows: list[dict[str, str]], fields: list[str]) -> None:
    if path.exists():
        raise FileExistsError(f"Refusing to overwrite: {path}")
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--tcmsp-compounds", type=Path, required=True)
    parser.add_argument("--swiss-targets", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    args = parser.parse_args()

    args.output_dir.mkdir(parents=True, exist_ok=False)
    compound_rows = read_rows(args.tcmsp_compounds)
    target_rows = read_rows(args.swiss_targets)

    compounds = [
        {
            "compound_id": row["MOL"],
            "compound_name": row["NAME"],
            "herb_source_archived": row["SOURCE"],
            "compound_source": "TCMSP",
        }
        for row in compound_rows
    ]
    predictions = [
        {
            "herb_source_archived": row["drug"],
            "compound_id": row["MOLID"],
            "compound_name": row["moleculename"],
            "predicted_target_label": row["GeneName"],
            "target_prediction_source": "SwissTargetPrediction",
        }
        for row in target_rows
    ]

    unique_predictions = {
        (
            row["herb_source_archived"],
            row["compound_id"],
            row["compound_name"],
            row["predicted_target_label"],
        )
        for row in predictions
    }
    compound_ids = {row["compound_id"] for row in compounds}
    predicted_compound_ids = {row["compound_id"] for row in predictions}
    herb_counts = Counter(row["herb_source_archived"] for row in predictions)

    summary = [
        {"metric": "archived_TCMSP_compound_rows", "value": len(compounds)},
        {"metric": "unique_TCMSP_compound_ids", "value": len(compound_ids)},
        {"metric": "archived_SwissTargetPrediction_rows", "value": len(predictions)},
        {"metric": "unique_herb_compound_target_relations", "value": len(unique_predictions)},
        {"metric": "compounds_with_retained_prediction_rows", "value": len(predicted_compound_ids)},
        {"metric": "compounds_without_retained_prediction_rows", "value": len(compound_ids - predicted_compound_ids)},
        {"metric": "Cuscuta_chinensis_prediction_rows", "value": herb_counts["tusizi"]},
        {"metric": "Ligustrum_lucidum_prediction_rows", "value": herb_counts["nvzhenzi"]},
        {"metric": "Eclipta_prostrata_prediction_rows", "value": herb_counts["mohanlian"]},
        {"metric": "tcmsp_compound_source_sha256", "value": hashlib.sha256(args.tcmsp_compounds.read_bytes()).hexdigest().upper()},
        {"metric": "swiss_target_source_sha256", "value": hashlib.sha256(args.swiss_targets.read_bytes()).hexdigest().upper()},
    ]

    write_csv(
        args.output_dir / "TCMSP_active_compounds.csv",
        compounds,
        ["compound_id", "compound_name", "herb_source_archived", "compound_source"],
    )
    write_csv(
        args.output_dir / "SwissTargetPrediction_compound_target_rows.csv",
        predictions,
        [
            "herb_source_archived",
            "compound_id",
            "compound_name",
            "predicted_target_label",
            "target_prediction_source",
        ],
    )
    write_csv(args.output_dir / "compound_target_provenance_summary.csv", summary, ["metric", "value"])


if __name__ == "__main__":
    main()
