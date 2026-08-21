#!/usr/bin/env python3
"""Verify the exact current Human Genomics submission package.

The submission package remains outside this repository. This command checks
the separately held flat package against the current 19-file checksum table,
then applies format and content-structure checks that do not alter any file.
The archived v0.1.0 verifier remains available as ``verify_submission_assets.py``.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
from pathlib import Path
from zipfile import ZipFile


ROOT = Path(__file__).resolve().parents[2]
CHECKSUMS = ROOT / "CURRENT_MANUSCRIPT_ASSET_CHECKSUMS.tsv"
EXPECTED_FIGURES = {
    "Figure_1.png": ((4464, 4444), 600),
    "Figure_2.png": ((4466, 4868), 300),
    "Figure_3.png": ((4500, 4688), 600),
    "Figure_4.png": ((8184, 4358), 600),
    "Figure_5.png": ((3250, 3815), 300),
    "Figure_6.png": ((4323, 4260), 600),
    "Figure_7.png": ((5000, 2380), 600),
    "Figure_8.png": ((4322, 5314), 600),
}
EXPECTED_ROLES = {
    "manuscript": 1,
    "main_figure": 8,
    "main_table": 2,
    "graphical_abstract": 1,
    "cover_letter": 1,
    "additional_file": 6,
}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest().upper()


def nonempty_rows(worksheet) -> list[tuple[object, ...]]:
    return [
        tuple(row)
        for row in worksheet.iter_rows(values_only=True)
        if any(value is not None and str(value).strip() for value in row)
    ]


def verify_docx_package(path: Path) -> None:
    with ZipFile(path, "r") as archive:
        bad = archive.testzip()
        if bad is not None:
            raise ValueError(f"{path.name}: corrupt DOCX member {bad}")
        names = set(archive.namelist())
        if "word/document.xml" not in names:
            raise ValueError(f"{path.name}: missing word/document.xml")
        joined = b"\n".join(
            archive.read(name)
            for name in names
            if name.endswith(".xml") or name.endswith(".rels")
        )
        for forbidden in (b"contractReview", b"vas-ai-hub"):
            if forbidden in joined:
                raise ValueError(
                    f"{path.name}: hidden WPS review store detected ({forbidden.decode()})"
                )


def verify_images(package_root: Path) -> None:
    try:
        from PIL import Image
    except ImportError as exc:
        raise RuntimeError("Pillow is required to verify figure dimensions") from exc

    for filename, (expected_size, expected_dpi) in EXPECTED_FIGURES.items():
        path = package_root / filename
        with Image.open(path) as image:
            image.verify()
        with Image.open(path) as image:
            if image.size != expected_size:
                raise ValueError(
                    f"{filename}: expected {expected_size}, found {image.size}"
                )
            if image.mode not in {"RGB", "RGBA"}:
                raise ValueError(f"{filename}: unexpected color mode {image.mode}")
            dpi = image.info.get("dpi")
            if dpi is None or any(abs(value - expected_dpi) > 1.0 for value in dpi):
                raise ValueError(
                    f"{filename}: expected approximately {expected_dpi} dpi, found {dpi}"
                )
        if path.stat().st_size >= 10 * 1024 * 1024:
            raise ValueError(f"{filename}: exceeds the 10 MiB figure ceiling")

    graphical = package_root / "Graphical_Abstract.png"
    with Image.open(graphical) as image:
        image.verify()
    with Image.open(graphical) as image:
        if image.size != (920, 300):
            raise ValueError(
                f"Graphical_Abstract.png: expected (920, 300), found {image.size}"
            )


def verify_pdf(package_root: Path) -> None:
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise RuntimeError("pypdf is required to verify the supplementary PDF") from exc

    path = package_root / "Additional_file_2_Supplementary_Figures_S1-S16.pdf"
    reader = PdfReader(path)
    if reader.is_encrypted:
        raise ValueError(f"{path.name}: encrypted supplementary PDF")
    if len(reader.pages) != 16:
        raise ValueError(f"{path.name}: expected 16 pages, found {len(reader.pages)}")
    metadata = " ".join(str(value) for value in (reader.metadata or {}).values())
    for forbidden in ("S1-S19", "frozen source", "legend-margin correction"):
        if forbidden.lower() in metadata.lower():
            raise ValueError(f"{path.name}: stale/internal metadata contains {forbidden!r}")


def verify_workbooks(package_root: Path) -> list[str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for --inspect-workbooks") from exc

    results: list[str] = []
    file1 = package_root / "Additional_file_1_Supplementary_Tables_S1-S6.xlsx"
    workbook = load_workbook(file1, read_only=True, data_only=True)
    expected_rows = {
        "S1_prescription_records": 391,
        "S3_association_rules": 13,
        "S4_candidate_pool_126": 127,
        "S4_SwissTarget_predictions": 2358,
        "S5_priority_candidates": 31,
    }
    for sheet, expected in expected_rows.items():
        observed = len(nonempty_rows(workbook[sheet]))
        if observed != expected:
            raise ValueError(f"{sheet}: expected {expected} nonempty rows, found {observed}")
        results.append(f"{sheet}={observed - 1} data rows")
    workbook.close()

    file3 = package_root / "Additional_file_3_Supplementary_Table_S7.xlsx"
    workbook = load_workbook(file3, read_only=True, data_only=True)
    all_docking = len(nonempty_rows(workbook["All docking"])) - 1
    if all_docking != 138:
        raise ValueError(f"All docking: expected 138 runs, found {all_docking}")
    workbook.close()
    results.append("All docking=138 data rows")

    for filename in (
        "Additional_file_4_Supplementary_Table_S8.xlsx",
        "Additional_file_5_Supplementary_Table_S10.xlsx",
        "Additional_file_6_Supplementary_Table_S9.xlsx",
    ):
        workbook = load_workbook(package_root / filename, read_only=True, data_only=True)
        if not workbook.sheetnames:
            raise ValueError(f"{filename}: no worksheets")
        results.append(f"{filename}={len(workbook.sheetnames)} worksheets")
        workbook.close()
    return results


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("package_root", type=Path)
    parser.add_argument("--inspect-workbooks", action="store_true")
    args = parser.parse_args()
    package_root = args.package_root.resolve()
    if not package_root.is_dir():
        raise SystemExit(f"Submission package is not a directory: {package_root}")

    with CHECKSUMS.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle, delimiter="\t"))
    filenames = [row["filename"] for row in rows]
    if len(rows) != 19 or len(filenames) != len(set(filenames)):
        raise SystemExit("Current checksum table must contain 19 unique assets")
    role_counts = {
        role: sum(row["package_role"] == role for row in rows)
        for role in EXPECTED_ROLES
    }
    if role_counts != EXPECTED_ROLES:
        raise SystemExit(f"Unexpected current asset roles: {role_counts}")

    actual_files = {path.name for path in package_root.iterdir() if path.is_file()}
    expected_files = set(filenames)
    if actual_files != expected_files:
        missing = sorted(expected_files - actual_files)
        extra = sorted(actual_files - expected_files)
        raise SystemExit(f"Current package coverage mismatch: missing={missing}, extra={extra}")

    failures: list[str] = []
    for row in rows:
        path = package_root / row["filename"]
        observed_size = path.stat().st_size
        observed_sha = sha256(path)
        if observed_size != int(row["bytes"]):
            failures.append(
                f"{path.name}: byte mismatch {observed_size} != {row['bytes']}"
            )
        if observed_sha != row["sha256"]:
            failures.append(
                f"{path.name}: SHA-256 mismatch {observed_sha} != {row['sha256']}"
            )
        if row["package_role"] == "additional_file" and observed_size >= 20_000_000:
            failures.append(f"{path.name}: exceeds the 20,000,000-byte additional-file ceiling")
    if failures:
        raise SystemExit("\n".join(failures))

    verify_images(package_root)
    verify_pdf(package_root)
    for filename in (
        "Manuscript.docx",
        "Cover_Letter.docx",
        "Table_1.docx",
        "Table_2.docx",
    ):
        verify_docx_package(package_root / filename)
    workbook_results = verify_workbooks(package_root) if args.inspect_workbooks else []

    print("PASS: exact 19-file current Human Genomics submission package verified")
    print("PASS: eight main figures, graphical abstract and 16-page supplementary PDF verified")
    if workbook_results:
        print("PASS: " + "; ".join(workbook_results))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
