#!/usr/bin/env python3
"""Verify a separately held Scientific Reports transfer package.

The package is never copied into this repository. The verifier is intentionally
strict: it accepts only the 16 planned transfer files, checks exact hashes from
CURRENT_MANUSCRIPT_ASSET_CHECKSUMS.tsv, audits manuscript supplementary first
use, and rejects prior-journal extras. Until the checksum table is refreshed
after final visual approval, this command exits with an actionable error rather
than treating the prior 19-file inventory as current.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import re
import xml.etree.ElementTree as ET
from collections import Counter
from pathlib import Path
from zipfile import ZipFile


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_CHECKSUMS = ROOT / "CURRENT_MANUSCRIPT_ASSET_CHECKSUMS.tsv"
W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"

EXPECTED_FILES = {
    "Manuscript.docx",
    "Cover_Letter.docx",
    *(f"Figure_{index}.png" for index in range(1, 9)),
    "Supplementary_Information.pdf",
    "Supplementary_Tables_S1-S6.xlsx",
    *(f"Supplementary_Table_S{index}.xlsx" for index in range(7, 11)),
}
EXPECTED_ROLES = {
    "manuscript": 1,
    "cover_letter": 1,
    "main_figure": 8,
    "supplementary_information": 1,
    "supplementary_table": 5,
}
EXPECTED_FIGURES = {
    "Figure_1.png": ((4464, 4444), 600),
    "Figure_2.png": ((4466, 4868), 300),
    "Figure_3.png": ((4500, 4688), 600),
    "Figure_4.png": ((8184, 4358), 600),
    "Figure_5.png": ((3250, 3815), 300),
    "Figure_6.png": ((4323, 4260), 600),
    "Figure_7.png": ((5000, 2380), 600),
    "Figure_8.png": ((4322, 5314), 600),
}
FORBIDDEN_FILENAMES = {
    "Graphical_Abstract.png",
    "Table_1.docx",
    "Table_2.docx",
}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest().upper()


def docx_text(path: Path) -> tuple[str, list[str]]:
    with ZipFile(path, "r") as archive:
        bad = archive.testzip()
        if bad is not None:
            raise ValueError(f"{path.name}: corrupt DOCX member {bad}")
        names = archive.namelist()
        if "word/document.xml" not in names:
            raise ValueError(f"{path.name}: missing word/document.xml")
        joined_xml = b"\n".join(
            archive.read(name)
            for name in names
            if name.endswith(".xml") or name.endswith(".rels")
        )
        for forbidden in (b"contractReview", b"vas-ai-hub"):
            if forbidden in joined_xml:
                raise ValueError(
                    f"{path.name}: hidden WPS review store detected "
                    f"({forbidden.decode()})"
                )
        root = ET.fromstring(archive.read("word/document.xml"))
        paragraphs: list[str] = []
        for paragraph in root.iter(f"{{{W_NS}}}p"):
            value = "".join(
                node.text or "" for node in paragraph.iter(f"{{{W_NS}}}t")
            ).strip()
            if value:
                paragraphs.append(value)
        media = sorted(name for name in names if name.startswith("word/media/"))
    return "\n".join(paragraphs), media


def first_unique(values: list[int]) -> list[int]:
    seen: set[int] = set()
    ordered: list[int] = []
    for value in values:
        if value not in seen:
            seen.add(value)
            ordered.append(value)
    return ordered


def supplement_numbers(text: str, kind: str) -> list[int]:
    if kind == "figure":
        prefix_pattern = re.compile(
            r"Supplementary\s+Fig(?:ure)?s?\.?\s*", re.IGNORECASE
        )
    else:
        prefix_pattern = re.compile(r"Supplementary\s+Tables?\s*", re.IGNORECASE)

    values: list[int] = []
    for prefix in prefix_pattern.finditer(text):
        tail = text[prefix.end() : prefix.end() + 120]
        stop_positions = [
            position
            for delimiter in (";", ".", ")", "\n")
            if (position := tail.find(delimiter)) >= 0
        ]
        chunk = tail[: min(stop_positions)] if stop_positions else tail
        for match in re.finditer(r"S(\d+)(?:\s*[–-]\s*S?(\d+))?", chunk):
            start = int(match.group(1))
            end = int(match.group(2)) if match.group(2) else start
            if end < start:
                raise ValueError(f"Descending supplementary range S{start}-S{end}")
            values.extend(range(start, end + 1))
    return values


def verify_manuscript(path: Path) -> None:
    text, media = docx_text(path)
    if media:
        raise ValueError(
            f"{path.name}: Scientific Reports text-only manuscript contains "
            f"embedded media: {media}"
        )
    lower = text.lower()
    for forbidden in ("additional file", "graphical abstract", "human genomics"):
        if forbidden in lower:
            raise ValueError(f"{path.name}: stale term present: {forbidden!r}")
    for match in re.finditer(
        r"\b(?:(?:Figure|Fig\.?)s?|Tables?)\s+S\d+", text
    ):
        prefix = text[max(0, match.start() - 24) : match.start()].rstrip().lower()
        if not prefix.endswith("supplementary"):
            raise ValueError(
                f"{path.name}: supplementary item lacks the 'Supplementary' prefix"
            )
    if re.search(
        r"Supplementary\s+Fig(?:ure)?s?\.?\s*S\d+[a-z]\b", text, re.IGNORECASE
    ):
        raise ValueError(f"{path.name}: supplementary panel-level citation found")
    if re.search(r"\b(?:Figure|Fig\.?)\s*\d+\s*[A-Z]\b", text):
        raise ValueError(f"{path.name}: uppercase main-panel citation found")

    figure_order = first_unique(supplement_numbers(text, "figure"))
    table_order = first_unique(supplement_numbers(text, "table"))
    if figure_order != list(range(1, 17)):
        raise ValueError(
            "Supplementary-figure first-use order mismatch: "
            f"expected S1-S16, found {figure_order}"
        )
    if table_order != list(range(1, 11)):
        raise ValueError(
            "Supplementary-table first-use order mismatch: "
            f"expected S1-S10, found {table_order}"
        )


def verify_cover_letter(path: Path) -> None:
    text, _ = docx_text(path)
    if "Scientific Reports" not in text:
        raise ValueError(f"{path.name}: journal name 'Scientific Reports' missing")
    lower = text.lower()
    for forbidden in ("human genomics", "bmc genomics", "additional file"):
        if forbidden in lower:
            raise ValueError(f"{path.name}: stale journal/package term {forbidden!r}")


def verify_images(package_root: Path) -> None:
    try:
        from PIL import Image
    except ImportError as exc:
        raise RuntimeError("Pillow is required to verify figure metadata") from exc

    maximum_print_width_inches = 180 / 25.4
    for filename, (expected_size, _legacy_density_tag) in EXPECTED_FIGURES.items():
        path = package_root / filename
        with Image.open(path) as image:
            image.verify()
        with Image.open(path) as image:
            if image.size != expected_size:
                raise ValueError(
                    f"{filename}: expected {expected_size}, found {image.size}"
                )
            if image.mode not in {"RGB", "RGBA"}:
                raise ValueError(f"{filename}: unexpected color mode {image.mode}")
            effective_dpi = image.size[0] / maximum_print_width_inches
            if effective_dpi < 300:
                raise ValueError(
                    f"{filename}: effective resolution at 180 mm width is only "
                    f"{effective_dpi:.1f} dpi"
                )
        if path.stat().st_size >= 10 * 1024 * 1024:
            raise ValueError(f"{filename}: exceeds the 10 MiB figure ceiling")


def verify_pdf(package_root: Path) -> None:
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise RuntimeError("pypdf is required to verify the supplement") from exc

    path = package_root / "Supplementary_Information.pdf"
    if path.stat().st_size >= 50 * 1024 * 1024:
        raise ValueError(f"{path.name}: exceeds 50 MiB")
    reader = PdfReader(path)
    if reader.is_encrypted:
        raise ValueError(f"{path.name}: encrypted PDF")
    if len(reader.pages) < 17:
        raise ValueError(
            f"{path.name}: expected cover/legend material plus 16 figures; "
            f"found only {len(reader.pages)} pages"
        )
    extracted = "\n".join(page.extract_text() or "" for page in reader.pages)
    lower = extracted.lower()
    for forbidden in ("additional file", "human genomics"):
        if forbidden in lower:
            raise ValueError(f"{path.name}: stale term present: {forbidden!r}")
    for index in range(1, 17):
        if not re.search(
            rf"Supplementary\s+Fig(?:ure)?s?\.?\s*S{index}\b",
            extracted,
            re.IGNORECASE,
        ):
            raise ValueError(f"{path.name}: Supplementary Fig. S{index} missing")


def nonempty_rows(worksheet) -> list[tuple[object, ...]]:
    return [
        tuple(row)
        for row in worksheet.iter_rows(values_only=True)
        if any(value is not None and str(value).strip() for value in row)
    ]


def workbook_error_cells(workbook) -> list[str]:
    failures: list[str] = []
    error_tokens = {"#REF!", "#DIV/0!", "#NAME?", "#VALUE!", "#N/A"}
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip() in error_tokens:
                    failures.append(f"{worksheet.title}!{cell.coordinate}={cell.value}")
    return failures


def verify_workbooks(package_root: Path) -> list[str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for --inspect-workbooks") from exc

    results: list[str] = []
    combined = package_root / "Supplementary_Tables_S1-S6.xlsx"
    workbook = load_workbook(combined, read_only=True, data_only=False)
    expected_rows = {
        "S1_prescription_records": 391,
        "S3_association_rules": 13,
        "S4_candidate_pool_126": 127,
        "S4_SwissTarget_predictions": 2358,
        "S5_priority_candidates": 31,
    }
    for sheet, expected in expected_rows.items():
        if sheet not in workbook.sheetnames:
            raise ValueError(f"{combined.name}: missing sheet {sheet}")
        observed = len(nonempty_rows(workbook[sheet]))
        if observed != expected:
            raise ValueError(f"{sheet}: expected {expected} rows, found {observed}")
        results.append(f"{sheet}={observed - 1} data rows")
    errors = workbook_error_cells(workbook)
    workbook.close()
    if errors:
        raise ValueError(f"{combined.name}: spreadsheet errors: {errors[:10]}")

    geneformer = package_root / "Supplementary_Table_S7.xlsx"
    workbook = load_workbook(geneformer, read_only=True, data_only=False)
    required_geneformer = {
        "README",
        "Candidate summary",
        "Donor effects",
        "LODO stability",
        "Bidirectional summary",
        "Observed recovery",
        "Cross-model",
    }
    missing = sorted(required_geneformer - set(workbook.sheetnames))
    errors = workbook_error_cells(workbook)
    sheet_count = len(workbook.sheetnames)
    workbook.close()
    if missing:
        raise ValueError(f"{geneformer.name}: missing sheets {missing}")
    if errors:
        raise ValueError(f"{geneformer.name}: spreadsheet errors: {errors[:10]}")
    results.append(f"{geneformer.name}={sheet_count} worksheets")

    sctenifold = package_root / "Supplementary_Table_S8.xlsx"
    workbook = load_workbook(sctenifold, read_only=True, data_only=False)
    if not workbook.sheetnames:
        raise ValueError(f"{sctenifold.name}: no worksheets")
    errors = workbook_error_cells(workbook)
    sheet_count = len(workbook.sheetnames)
    workbook.close()
    if errors:
        raise ValueError(f"{sctenifold.name}: spreadsheet errors: {errors[:10]}")
    results.append(f"{sctenifold.name}={sheet_count} worksheets")

    docking = package_root / "Supplementary_Table_S9.xlsx"
    workbook = load_workbook(docking, read_only=True, data_only=False)
    if "All docking" not in workbook.sheetnames:
        raise ValueError(f"{docking.name}: missing 'All docking' sheet")
    all_docking = len(nonempty_rows(workbook["All docking"])) - 1
    errors = workbook_error_cells(workbook)
    workbook.close()
    if all_docking != 138:
        raise ValueError(f"All docking: expected 138 runs, found {all_docking}")
    if errors:
        raise ValueError(f"{docking.name}: spreadsheet errors: {errors[:10]}")
    results.append("All docking=138 data rows")

    search = package_root / "Supplementary_Table_S10.xlsx"
    workbook = load_workbook(search, read_only=True, data_only=False)
    if "Table S10" not in workbook.sheetnames:
        raise ValueError(f"{search.name}: missing 'Table S10' sheet")
    errors = workbook_error_cells(workbook)
    sheet_count = len(workbook.sheetnames)
    workbook.close()
    if errors:
        raise ValueError(f"{search.name}: spreadsheet errors: {errors[:10]}")
    results.append(f"{search.name}={sheet_count} worksheets")
    return results


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("package_root", type=Path)
    parser.add_argument("--checksums", type=Path, default=DEFAULT_CHECKSUMS)
    parser.add_argument("--inspect-workbooks", action="store_true")
    args = parser.parse_args()

    package_root = args.package_root.resolve()
    if not package_root.is_dir():
        raise SystemExit(f"Submission package is not a directory: {package_root}")
    checksum_path = args.checksums.resolve()
    with checksum_path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle, delimiter="\t"))
    filenames = [row.get("filename", "") for row in rows]
    roles = Counter(row.get("package_role", "") for row in rows)
    if len(rows) != 16 or set(filenames) != EXPECTED_FILES:
        missing = sorted(EXPECTED_FILES - set(filenames))
        extra = sorted(set(filenames) - EXPECTED_FILES)
        raise SystemExit(
            "Scientific Reports checksum inventory is not final: "
            f"expected 16 rows; found {len(rows)}; missing={missing}; extra={extra}. "
            "Generate and visually approve the final package before replacing "
            "CURRENT_MANUSCRIPT_ASSET_CHECKSUMS.tsv."
        )
    if dict(roles) != EXPECTED_ROLES:
        raise SystemExit(f"Unexpected Scientific Reports asset roles: {dict(roles)}")
    if any(
        len(row.get("sha256", "")) != 64 or int(row.get("bytes", "0")) <= 0
        for row in rows
    ):
        raise SystemExit("Invalid Scientific Reports size or SHA-256 entry")

    actual_files = {path.name for path in package_root.iterdir() if path.is_file()}
    if actual_files & FORBIDDEN_FILENAMES or any(
        name.lower().startswith("additional_file") for name in actual_files
    ):
        raise SystemExit(
            "Forbidden prior-journal assets found: "
            f"{sorted((actual_files & FORBIDDEN_FILENAMES) | {name for name in actual_files if name.lower().startswith('additional_file')})}"
        )
    if actual_files != EXPECTED_FILES:
        missing = sorted(EXPECTED_FILES - actual_files)
        extra = sorted(actual_files - EXPECTED_FILES)
        raise SystemExit(f"Package coverage mismatch: missing={missing}, extra={extra}")

    failures: list[str] = []
    for row in rows:
        path = package_root / row["filename"]
        observed_size = path.stat().st_size
        observed_sha = sha256(path)
        if observed_size != int(row["bytes"]):
            failures.append(
                f"{path.name}: byte mismatch {observed_size} != {row['bytes']}"
            )
        if observed_sha != row["sha256"].upper():
            failures.append(
                f"{path.name}: SHA-256 mismatch {observed_sha} != {row['sha256']}"
            )
    if failures:
        raise SystemExit("\n".join(failures))

    verify_manuscript(package_root / "Manuscript.docx")
    verify_cover_letter(package_root / "Cover_Letter.docx")
    verify_images(package_root)
    verify_pdf(package_root)
    workbook_results = verify_workbooks(package_root) if args.inspect_workbooks else []

    print("PASS: exact 16-file Scientific Reports transfer package verified")
    print("PASS: strict Supplementary Fig. S1-S16 and Table S1-S10 first-use order")
    print("PASS: prior graphical abstract, standalone main tables and Additional files absent")
    if workbook_results:
        print("PASS: " + "; ".join(workbook_results))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
