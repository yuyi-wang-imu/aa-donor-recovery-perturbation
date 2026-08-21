#!/usr/bin/env python3
"""Verify the archived v0.1.0 publication figures and supplementary workbooks.

The archived package is supplied separately and is never copied into the
repository by this command. Exact sizes and SHA-256 values are checked first.
Both a flat directory and the earlier nested layout are supported. Use
``verify_current_submission_assets.py`` for the current eight-figure
Human Genomics package.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
from pathlib import Path


ROOT = Path(__file__).resolve().parents[2]
CHECKSUMS = ROOT / "PUBLICATION_ASSET_CHECKSUMS.tsv"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest().upper()


def nonempty_rows(worksheet) -> list[tuple[object, ...]]:
    return [
        tuple(row)
        for row in worksheet.iter_rows(values_only=True)
        if any(value is not None and str(value).strip() for value in row)
    ]


def resolve_submission_asset(package_root: Path, relative_path: str) -> Path:
    """Resolve one expected asset without weakening filename checks.

    The current submission package stores all submission assets directly under
    its root. Earlier packages used ``02_Main_Figures`` and
    ``04_Additional_Files``. The checksum table retains the earlier relative
    paths for provenance, while this resolver accepts either layout.
    """

    nested = package_root / Path(relative_path)
    flat = package_root / Path(relative_path).name
    matches = [candidate for candidate in (nested, flat) if candidate.is_file()]
    if len(matches) == 1:
        return matches[0]
    if len(matches) > 1:
        raise ValueError(
            f"ambiguous asset present in both flat and nested layouts: {relative_path}"
        )
    return nested


def inspect_workbooks(package_root: Path) -> list[str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required for --inspect-workbooks; install environment/requirements.txt"
        ) from exc

    file1 = resolve_submission_asset(
        package_root,
        "04_Additional_Files/Additional_file_1_Supplementary_Tables_S1-S6_PublicationReady.xlsx",
    )
    workbook = load_workbook(file1, read_only=True, data_only=True)
    expected_rows = {
        "S1_prescription_records": 391,
        "S3_association_rules": 13,
        "S4_candidate_pool_126": 127,
        "S4_SwissTarget_predictions": 2358,
        "S5_priority_candidates": 31,
    }
    results = []
    for sheet, expected in expected_rows.items():
        observed = len(nonempty_rows(workbook[sheet]))
        if observed != expected:
            raise ValueError(f"{sheet}: expected {expected} nonempty rows, found {observed}")
        results.append(f"{sheet}={observed - 1} data rows")
    field_text = "\n".join(
        "\t".join("" if value is None else str(value) for value in row)
        for row in nonempty_rows(workbook["Field_dictionary"])
    )
    if "Twelve rules under the verified thresholds" not in field_text:
        raise ValueError("Field_dictionary does not contain the corrected twelve-rule wording")
    workbook.close()

    file3 = resolve_submission_asset(
        package_root,
        "04_Additional_Files/Additional_file_3_Supplementary_Table_S7.xlsx",
    )
    workbook = load_workbook(file3, read_only=True, data_only=True)
    all_docking = len(nonempty_rows(workbook["All docking"])) - 1
    if all_docking != 138:
        raise ValueError(f"All docking: expected 138 runs, found {all_docking}")
    workbook.close()
    results.append("All docking=138 data rows")
    return results


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("package_root", type=Path)
    parser.add_argument("--inspect-workbooks", action="store_true")
    args = parser.parse_args()
    package_root = args.package_root.resolve()

    with CHECKSUMS.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle, delimiter="\t"))
    failures = []
    for row in rows:
        try:
            path = resolve_submission_asset(package_root, row["relative_submission_path"])
        except ValueError as exc:
            failures.append(str(exc))
            continue
        if not path.is_file():
            failures.append(f"missing: {row['relative_submission_path']}")
            continue
        if path.stat().st_size != int(row["bytes"]):
            failures.append(f"size mismatch: {row['relative_submission_path']}")
            continue
        if sha256(path) != row["sha256"]:
            failures.append(f"SHA-256 mismatch: {row['relative_submission_path']}")
    if failures:
        raise SystemExit("\n".join(failures))
    print(f"PASS: {len(rows)} submission assets match size and SHA-256")

    if args.inspect_workbooks:
        for message in inspect_workbooks(package_root):
            print("PASS:", message)
        print("PASS: corrected twelve-rule field dictionary wording")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
