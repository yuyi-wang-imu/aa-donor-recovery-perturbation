#!/usr/bin/env python3
"""Extract the 12 two-herb Apriori rules from the original SPSS workbook.

The selection reproduces the 316 manuscript method: SPSS Modeler 18.0,
minimum antecedent support 10%, minimum confidence 88%, and maximum
antecedent size 2.  The manuscript table is the subset with exactly two herbs
in the antecedent.  No source workbook is modified.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import pathlib

from openpyxl import load_workbook


LATIN = {
    "女贞子": "Ligustri Lucidi Fructus",
    "墨旱莲": "Ecliptae Herba",
    "菟丝子": "Cuscutae Semen",
    "当归": "Angelicae Sinensis Radix",
    "鹿角胶": "Cervi Cornus Colla",
    "白术": "Atractylodis Macrocephalae Rhizoma",
    "阿胶": "Asini Corii Colla",
    "熟地黄": "Rehmanniae Radix Praeparata",
    "仙鹤草": "Agrimoniae Herba",
    "巴戟天": "Morindae Officinalis Radix",
    "淫羊藿": "Epimedii Folium",
    "党参": "Codonopsis Radix",
    "茯苓": "Poria",
    "黄芪": "Astragali Radix",
    "白芍": "Paeoniae Radix Alba",
    "黄精": "Polygonati Rhizoma",
}


def sha256(path: pathlib.Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=pathlib.Path, required=True)
    parser.add_argument("--json", type=pathlib.Path, required=True)
    parser.add_argument("--csv", type=pathlib.Path, required=True)
    args = parser.parse_args()
    for output in (args.json, args.csv):
        if output.exists():
            raise SystemExit(f"Refusing to overwrite existing output: {output}")

    wb = load_workbook(args.source, data_only=True, read_only=True)
    ws = wb.worksheets[1]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    all_rows = []
    selected = []
    for source_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        record = dict(zip(headers, row))
        record["source_row"] = source_row
        all_rows.append(record)
        antecedent = record["前项"]
        support = float(record["支持度百分比"])
        confidence = float(record["置信度百分比"])
        parts = [part.strip() for part in antecedent.split(" and ")] if isinstance(antecedent, str) else []
        if len(parts) != 2 or support < 10 or confidence < 88:
            continue
        missing = [name for name in parts + [record["后项"]] if name not in LATIN]
        if missing:
            raise RuntimeError(f"Missing Latin mapping: {missing}")
        joint_n = round(float(record["规则支持度百分比"]) * 390 / 100)
        selected.append(
            {
                "rank": len(selected) + 1,
                "antecedent_1_zh": parts[0],
                "antecedent_1_en": LATIN[parts[0]],
                "antecedent_2_zh": parts[1],
                "antecedent_2_en": LATIN[parts[1]],
                "consequent_zh": record["后项"],
                "consequent_en": LATIN[record["后项"]],
                "rule_en": f"{LATIN[parts[0]]} + {LATIN[parts[1]]} → {LATIN[record['后项']]}",
                "spss_rule_id": int(record["规则标识"]),
                "antecedent_n": int(record["实例"]),
                "joint_n": int(joint_n),
                "antecedent_coverage_pct": float(record["支持度百分比"]),
                "joint_support_pct": float(record["规则支持度百分比"]),
                "confidence_pct": float(record["置信度百分比"]),
                "lift": float(record["增益"]),
                "deployment_ability_pct": float(record["部署能力"]),
                "source_row": source_row,
            }
        )
    wb.close()

    if len(selected) != 12:
        raise RuntimeError(f"Expected 12 two-herb rules, found {len(selected)}")
    core = selected[0]
    expected = {
        "antecedent_n": 45,
        "joint_n": 44,
        "antecedent_coverage_pct": 45 / 390 * 100,
        "joint_support_pct": 44 / 390 * 100,
        "confidence_pct": 44 / 45 * 100,
    }
    for key, value in expected.items():
        if abs(core[key] - value) > 1e-9:
            raise RuntimeError(f"Core rule mismatch for {key}: {core[key]} vs {value}")

    payload = {
        "source_file": args.source.name,
        "source_sha256": sha256(args.source),
        "source_sheet": ws.title,
        "source_output_rule_count": len(all_rows),
        "selection": {
            "software": "IBM SPSS Modeler 18.0",
            "algorithm": "Apriori",
            "minimum_antecedent_support_pct": 10,
            "minimum_confidence_pct": 88,
            "maximum_antecedent_size": 2,
            "manuscript_table_antecedent_size": 2,
        },
        "network_visualization": {"weak_link_upper_limit": 15, "strong_link_lower_limit": 35},
        "hierarchical_clustering": {
            "software": "IBM SPSS Statistics 26.0",
            "linkage": "between-groups linkage",
            "minimum_clusters": 2,
        },
        "selected_rule_count": len(selected),
        "rules": selected,
    }
    args.json.parent.mkdir(parents=True, exist_ok=True)
    args.csv.parent.mkdir(parents=True, exist_ok=True)
    args.json.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    fields = list(selected[0])
    with args.csv.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(selected)
    print(f"PASS rules={len(selected)} core={core['rule_en']}")
    print(f"JSON {args.json}")
    print(f"CSV {args.csv}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
